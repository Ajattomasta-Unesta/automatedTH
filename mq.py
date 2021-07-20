from openpyxl import load_workbook


def load_data(filepath) :
    wb = load_workbook(filepath, data_only=True) #filepath
    #load_ws = load_wb['Sheet1']

    ws = wb[wb.sheetnames[0]]

    print("A")
    all_values = []
    for row in ws.rows:
        row_value = []
        for cell in row:
            row_value.append(cell.value)
        all_values.append(row_value)
    print(all_values)

    #L1-Primary infos
    L1 = all_values[1]
    print(L1)

    #L2-Default infos
    L2 = all_values[3][:5]
    print(L2)

    #L3-LIST
    L3 = []
    for row in ws.iter_rows(min_row=6):
        if row[0].value == None : break
        row_value = []
        for cell in row[0:3]:
            row_value.append(cell.value)
        L3.append(row_value)
    print(L3)

    #DATA INSERT
    bt = "O"
    btp = "  "
    if int(L1[0]) == 2 : bt, btp = btp, bt

    e1 = "OO❖"
    #if L1[8] != None or not L1[8].isspace() : e1 = L1[8]
    print(e1)

    b1, b2 = "", ""
    try: b1 = L2[2].date()
    except : pass

    try: b2 = L2[3].date()
    except : pass

    pddata = {'x1' : bt, 'x2' : btp, 'const' : L1[6],
              'a1' : L1[1],
              'a2' : L1[2],
              'a3' : L1[3],
              'a4' : L1[4],
              'a5' : L2[0],
              'a6' : L2[1],
              'b1' : b1,
              'b2' : b2,
              'b3' : L1[5],
              'd1' : L2[4].year,
              'd2' : L2[4].month,
              'd3' : L2[4].day,
              'd4' : L1[7]
              }

    ppdata = L3

    return pddata, ppdata

print(load_data("xl.xlsx"))