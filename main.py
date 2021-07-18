import time
from tkinter import *
from tkinter import filedialog
import shutil
import math
from openpyxl import load_workbook
import win32com.client as win32
import pyperclip
from sys import exit
from getmac import get_mac_address


def load_data(filepath) :
    wb = load_workbook(filepath, data_only=True) #filepath
    #load_ws = load_wb['Sheet1']

    ws = wb[wb.sheetnames[0]]

    print("A")
    all_values = []
    for row in ws.rows:
        row_value = []
        for cell in row:
            row_value.append(cell.value)
        all_values.append(row_value)
    print(all_values)

    #L1-Primary infos
    L1 = all_values[1]
    print(L1)

    #L2-Default infos
    L2 = all_values[3][:5]
    print(L2)

    #L3-LIST
    L3 = []
    for row in ws.iter_rows(min_row=6):
        if row[0].value == None : break
        row_value = []
        for cell in row[0:3]:
            row_value.append(cell.value)
        L3.append(row_value)
    print(L3)

    #DATA INSERT
    bt = "O"
    btp = "  "
    if int(L1[0]) == 2 : bt, btp = btp, bt

    pddata = {'x1' : bt, 'x2' : btp, 'const' : L1[6],
              'a1' : L1[1],
              'a2' : L1[2],
              'a3' : L1[3],
              'a4' : L1[4],
              'a5' : L2[0],
              'a6' : L2[1],
              'b1' : L2[2].date(),
              'b2' : L2[3].date(),
              'b3' : L1[5],
              'd1' : L2[4].year,
              'd2' : L2[4].month,
              'd3' : L2[4].day,
              'd4' : L1[7]
              }
    ppdata = L3
    print(L2[3].date())
    return pddata, ppdata

def inputhwp(path, pddata, ppdata) :
    hwp = win32.Dispatch("HWPFrame.HwpObject")

    #hwp.RegisterModule("FilePathCheckDLL", "FilePathCheckerModule")
    hwp.XHwpWindows.Item(0).Visible = True

    now = time.localtime()
    tm = "%04d-%02d-%02d-%02d-%02d-%02d" % (now.tm_year, now.tm_mon, now.tm_mday, now.tm_hour, now.tm_min, now.tm_sec)
    filepath = path+"/경비원배치폐지신고서_"+tm+".hwp"

    shutil.copyfile("form1.hwp", filepath)

    hwp.Open(filepath, "HWP", "forceopen:true")
    '''hwp.PutFieldText("x1", "  ")
    hwp.PutFieldText("x2", "O")
    hwp.PutFieldText("a4", "010-9999-9999")'''

    try :
        fl = hwp.GetFieldList(0,2)
    except:pass

    field_list = [i for i in fl.split("\x02")]
    print(field_list)

    # 개수따라 페이지생성
    # 파일생성 - 기본정보 - 4개이상일경우 int(4나누기+소수점 올림) -1개만큼큼 페이지 추가 - 페이지별로 사람정보, 페이지지정
    pgcnt = math.ceil(len(ppdata)/4)
    print("PGCNT : ", pgcnt)

    hwp.MovePos(2, None, None)
    hwp.Run('SelectAll')  # Ctrl-A (전체선택)
    hwp.Run('Copy')  # Ctrl-C (복사)
    hwp.MovePos(3, None, None)  # 문서 끝으로 이동

    for x in range(pgcnt-1) :
        try:
            hwp.Run('Paste')  # Ctrl-V (붙여넣기)
            print("RUN PASTE")
        except:pass

        try:
            hwp.MovePos(3, None, None)  # 문서 끝으로 이동
            print("RUN MOVEPOS 3")
        except : pass

    hwp.Run('Copy')
        # 개수따라 페이지생성
######################################################
    tmp = [['테스트', 1234567, 765544321], ['테스트1', 21234567, 1765544321]]

    idx = 1
    pg = 0
    fdata = []

    for i in ppdata:  # ppdata
        if idx > 4:
            pg = pg + 1
            idx = 1

        # f.append("C"+str(idx)+"1")
        # d.append("CONST") #pddata'const'

        fdata.append(("c" + str(idx) + "1", pg, pddata['const']))  # pddata'const'

        for j in range(2, 5):
            # f.append("C" + str(idx) + str(j))
            # d.append(i[j-2])
            fdata.append(("c" + str(idx) + str(j), pg, i[j - 2]))

        idx = idx + 1
    print(fdata)
#################################################################
    for x in field_list :
        try :
            hwp.PutFieldText(x, str(pddata[x]))
        except :
            print(x)

    for field, page, data in fdata :
        print(field, page, data, f'{field}{{{{{page}}}}}')
        hwp.PutFieldText(f'{field}{{{{{page}}}}}', str(data))

def load_xl():
    filename = filedialog.askopenfilename(initialdir="/", title="Select file", filetypes=(("Excel files", "*.xlsx"), ("all files", "*.*")))
    print(filename)
    stv_xl.set(filename)
    print(stv_xl.get())

def load_ph():
    dirname = filedialog.askdirectory(initialdir="/", title="Select file")
    print(dirname)
    stv_ph.set(dirname)
    print(stv_ph.get())

def execv():
    pd, pp = load_data(stv_xl.get())
    print("PP : ", pp)
    print("PD : ", pd)
    inputhwp(stv_ph.get(), pd, pp)

    pyperclip.copy("")

if __name__ == "__main__":

    ###################CERT####################
    try:
        ff = open("moonlight.dat", "r")
        if ff.read() == get_mac_address():
            print("VABENE")
        else:
            raise Exception
    except:
        exit()

    ff.close()
    ###################CERT####################

    root = Tk()
    root.title("경비원 배치폐지신고서 자동생성")

    lbl = Label(root, text="경비원 배치폐지신고서")
    lbl.grid(row=0, column=0, padx=15, pady=15)

    btn_xl = Button(root, text="엑셀 파일 선택", command = load_xl)
    btn_xl.grid(row=1, column=0, padx=15, pady=15)

    stv_xl = StringVar()
    lb_xl = Label(root, text="엑셀 파일 경로", textvariable=stv_xl)
    lb_xl.grid(row=1, column=1, padx=15, pady=15)

    btn_ph = Button(root, text="내보낼 폴더 선택", command=load_ph)
    btn_ph.grid(row=2, column=0, padx=15, pady=15)

    stv_ph = StringVar()
    lb_ph = Label(root, text="내보낼 폴더 경로", textvariable=stv_ph)
    lb_ph.grid(row=2, column=1, padx=15, pady=15)

    lb_xx = Label(root, text="파일명 입력란\n미입력시 이름 자동생성, 확장자 입력 필요 없음")
    lb_xx.grid(row=3, column=0, padx=25, pady=25)

    stv_fn = StringVar()
    sheetname = Entry(root, text="Sheet1", textvariable=stv_fn)
    sheetname.grid(row=3, column=1, padx=25, pady=25)

    btn_st = Button(root, text="생성 시작", command=execv)
    btn_st.grid(row=3, column=2, padx=25, pady=25)

    root.mainloop()
