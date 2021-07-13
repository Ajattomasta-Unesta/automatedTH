from openpyxl import load_workbook

wb = load_workbook("xl.xlsx", data_only=True)
#load_ws = load_wb['Sheet1']

ws = wb[wb.get_sheet_names()[0]]

all_values = []
for row in ws.rows:
    row_value = []
    for cell in row[0:3]:
        row_value.append(cell.value)
    all_values.append(row_value)
print(all_values)


#L1-Primary infos
l1_values = []
for row in ws.iter_rows(min_row=2, max_row=2):
    row_value = []
    for cell in row:
        row_value.append(cell.value)
    l1_values.append(row_value)
print(l1_values)

#L2-Default infos
l2_values = []
for row in ws.iter_rows(min_row=4, max_row=4):
    row_value = []
    for cell in row[0:5]:
        row_value.append(cell.value)
    l2_values.append(row_value)
print(l2_values)

#L3-LIST
l3_values = []
for row in ws.iter_rows(min_row=6):
    row_value = []
    for cell in row[0:3]:
        row_value.append(cell.value)
    l3_values.append(row_value)
print(l3_values)