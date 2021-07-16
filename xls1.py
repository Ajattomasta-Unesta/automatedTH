import math

from openpyxl import load_workbook
import win32com.client as win32

def load_data(filepath) :
    wb = load_workbook("xl.xlsx", data_only=True) #filepath
    #load_ws = load_wb['Sheet1']

    ws = wb[wb.sheetnames[0]]
    commonrail = []

    print("A")
    all_values = []
    for row in ws.rows:
        row_value = []
        for cell in row:
            row_value.append(cell.value)
        all_values.append(row_value)
    print(all_values)


    #L1-Primary infos
    print(all_values[1])

    #L2-Default infos
    print(all_values[3][:5])

    #L3-LIST
    l3_values = []
    for row in ws.iter_rows(min_row=6):
        if row[0].value == None : break
        row_value = []
        for cell in row[0:3]:
            row_value.append(cell.value)
        l3_values.append(row_value)
    print(l3_values)

    pddata=0
    ppdata=0
    return pddata, ppdata

def inputhwp(path, pddata, ppdata) :
    hwp = win32.Dispatch("HWPFrame.HwpObject")

    #hwp.RegisterModule("FilePathCheckDLL", "FilePathCheckerModule")
    hwp.XHwpWindows.Item(0).Visible = True

    hwp.Open("C:\\Users\\Luca\\PycharmProjects\\automatedTH\\form1.hwp")
    '''hwp.PutFieldText("x1", "  ")
    hwp.PutFieldText("x2", "O")
    hwp.PutFieldText("a4", "010-9999-9999")'''

    field_list = [i for i in hwp.GetFieldList().split("\x02")]
    print(field_list)

    # 개수따라 페이지생성
    # 파일생성 - 기본정보 - 4개이상일경우 int(4나누기+소수점 올림) -1개만큼큼 페이지 추가 - 페이지별로 사람정보, 페이지지정
    pgcnt = math.ceil(len(ppdata)/4)

    for x in range(pgcnt-1) :
        hwp.MovePos(2)
        hwp.Run('SelectAll')  # Ctrl-A (전체선택)
        hwp.Run('Copy')  # Ctrl-C (복사)
        hwp.MovePos(3)  # 문서 끝으로 이동

        #내용 붙여넣기
        for i in range(1):
            hwp.Run('Paste')  # Ctrl-V (붙여넣기)
            hwp.MovePos(3)  # 문서 끝으로 이동
        hwp.Run('Copy')
    # 개수따라 페이지생성

    tmp = [['테스트', 1234567, 765544321], ['테스트1', 21234567, 1765544321]]

    idx = 1
    f = []
    d = []#dict로수정
    for i in tmp :
        f.append("C"+str(idx)+"1")
        d.append("CONST")
        for j in range(2,5) :
            f.append("C" + str(idx) + str(j))
            d.append(i[j-2])
        idx = idx + 1
    print(f)
    print(d)

    for x in field_list :
        hwp.PutFieldText(x, str(pddata[x]))


load_data("a")
inputhwp("a", "pddata", "pp")